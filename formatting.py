import openpyxl
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

def adjust_dimensions(workbook):
    worksheet = workbook.active
    max_row = worksheet.max_row
    max_col = worksheet.max_column

    # adjust row heights
    for row in range(1, max_row + 1):
        if any(cell.value for cell in worksheet[row]):
            worksheet.row_dimensions[row].height = 45 # change this for size

    # adjust column widths
    for col in range(1, max_col + 1):
        column_letter = get_column_letter(col)
        if any(cell.value for cell in worksheet[column_letter]):
            worksheet.column_dimensions[column_letter].width = 13 # change this for size

    return workbook

def merge_cells(workbook):
    worksheet = workbook.active

    # title
    worksheet.merge_cells('A1:N1')
    worksheet['A1'].alignment = Alignment(horizontal='center', vertical='center')

    # header data
    worksheet.merge_cells('B2:C2')
    worksheet.merge_cells('E2:F2')
    worksheet.merge_cells('H2:I2')
    worksheet.merge_cells('K2:N2')
    worksheet.merge_cells('B3:C3')
    worksheet.merge_cells('E3:F3')
    worksheet.merge_cells('H3:I3')

    # for 8d and problem description
    merge_same_value_cells(worksheet, 1, 5)  # Column A
    merge_same_value_cells(worksheet, 2, 5)  # Column B

def merge_same_value_cells(worksheet, start_col, start_row):
    max_row = worksheet.max_row
    current_col = get_column_letter(start_col)
    
    merge_start = start_row
    current_value = worksheet[f"{current_col}{merge_start}"].value

    for row in range(start_row + 1, max_row + 1):
        cell_value = worksheet[f"{current_col}{row}"].value
        
        if cell_value != current_value:
            if row - merge_start > 1:
                worksheet.merge_cells(f"{current_col}{merge_start}:{current_col}{row - 1}")
            
            merge_start = row
            current_value = cell_value

    # Check if the last set of cells needs to be merged
    if max_row - merge_start >= 1:
        worksheet.merge_cells(f"{current_col}{merge_start}:{current_col}{max_row}")

def format_title(worksheet, cell):
    worksheet[cell].font = Font(size=16, name='Arial', bold=True)

def format_upper(worksheet, cell):
    worksheet[cell].fill = PatternFill(start_color='C6E0B4', fill_type='solid')
    worksheet[cell].font = Font(size=8, name='Arial', bold=True)
    worksheet[cell].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
def format_lower(worksheet, cell):
    worksheet[cell].fill = PatternFill(start_color='BDD7EE', fill_type='solid')
    worksheet[cell].font = Font(size=8, name='Arial', bold=True)
    worksheet[cell].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

def format_data(worksheet, cell):
    worksheet[cell].font = Font(size=8, name='Arial')
    worksheet[cell].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)